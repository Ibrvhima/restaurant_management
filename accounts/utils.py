from django.http import HttpResponse
from django.utils import timezone
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal
import io

def export_dashboard_excel(request):
    """Exporter les données du dashboard en format Excel"""
    # Récupérer les données
    from orders.models import Commande, EtatCommande
    from restaurant.models import Plat
    from payments.models import Paiement
    from expenses.models import Depense
    
    today = timezone.now().date()
    
    # Statistiques des commandes
    total_commandes = Commande.objects.count()
    commandes_en_cours = Commande.objects.filter(
        etat__in=[EtatCommande.EN_ATTENTE, EtatCommande.EN_PREPARATION, EtatCommande.EN_COURS]
    ).count()
    commandes_terminees = Commande.objects.filter(etat=EtatCommande.TERMINEE).count()
    
    # Statistiques des plats
    total_plats = Plat.objects.count()
    plats_disponibles = Plat.objects.filter(disponible=True).count()
    
    # Statistiques des paiements du jour
    paiements_aujourdhui = Paiement.objects.filter(date_paiement__date=today).count()
    total_ventes_aujourdhui = Paiement.objects.filter(date_paiement__date=today).aggregate(
        total=models.Sum('montant')
    )['total'] or 0
    
    # Statistiques des dépenses du jour
    depenses_aujourdhui = Depense.objects.filter(date_depense=today).count()
    total_depenses_aujourdhui = Depense.objects.filter(date_depense=today).aggregate(
        total=models.Sum('montant')
    )['total'] or 0
    
    # Solde de caisse
    total_entrées = Paiement.objects.aggregate(total=models.Sum('montant'))['total'] or 0
    total_sorties = Depense.objects.aggregate(total=models.Sum('montant'))['total'] or 0
    solde_caisse = total_entrées - total_sorties
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Restaurant"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # En-tête
    ws.merge_cells('A1:D1')
    ws['A1'] = f"RAPPORT DASHBOARD RESTAURANT - {today.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=16, color="2E86AB")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws['A2'].font = Font(size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Section Statistiques
    row = 4
    ws[f'A{row}'] = "STATISTIQUES GLOBALES"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="2E86AB")
    row += 1
    
    # En-têtes de tableau
    headers = ["Catégorie", "Indicateur", "Valeur", "Détails"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    row += 1
    
    # Données des commandes
    data = [
        ("Commandes", "Total", total_commandes, ""),
        ("", "En cours", commandes_en_cours, f"{(commandes_en_cours/total_commandes*100):.1f}%" if total_commandes > 0 else "0%"),
        ("", "Terminées", commandes_terminees, f"{(commandes_terminees/total_commandes*100):.1f}%" if total_commandes > 0 else "0%"),
        ("Plats", "Total", total_plats, ""),
        ("", "Disponibles", plats_disponibles, f"{(plats_disponibles/total_plats*100):.1f}%" if total_plats > 0 else "0%"),
        ("Paiements du jour", "Nombre", paiements_aujourdhui, ""),
        ("", "Montant total", f"{total_ventes_aujourdhui:,.0f} GNF", ""),
        ("Dépenses du jour", "Nombre", depenses_aujourdhui, ""),
        ("", "Montant total", f"{total_depenses_aujourdhui:,.0f} GNF", ""),
        ("Caisse", "Total entrées", f"{total_entrées:,.0f} GNF", ""),
        ("", "Total sorties", f"{total_sorties:,.0f} GNF", ""),
        ("", "Solde actuel", f"{solde_caisse:,.0f} GNF", "VERT" if solde_caisse >= 0 else "ROUGE"),
    ]
    
    for category, indicator, value, details in data:
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=indicator)
        ws.cell(row=row, column=3, value=value)
        ws.cell(row=row, column=4, value=details)
        
        # Couleur pour le solde
        if category == "Caisse" and indicator == "Solde actuel":
            if solde_caisse >= 0:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            else:
                ws.cell(row=row, column=3).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        # Bordures
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Section Transactions du jour
    row += 2
    ws[f'A{row}'] = "TRANSACTIONS DU JOUR"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="2E86AB")
    row += 1
    
    # Paiements du jour
    ws[f'A{row}'] = "Paiements"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="28A745")
    row += 1
    
    headers_paiements = ["ID", "Commande", "Montant", "Méthode", "Date"]
    for col, header in enumerate(headers_paiements, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        cell.border = border
        cell.alignment = center_alignment
    
    row += 1
    paiements_du_jour = Paiement.objects.filter(date_paiement__date=today).order_by('-date_paiement')
    for paiement in paiements_du_jour:
        ws.cell(row=row, column=1, value=paiement.id)
        ws.cell(row=row, column=2, value=f"#{paiement.commande.id}")
        ws.cell(row=row, column=3, value=f"{paiement.montant:,.0f}")
        ws.cell(row=row, column=4, value=paiement.get_methode_display())
        ws.cell(row=row, column=5, value=paiement.date_paiement.strftime('%H:%M:%S'))
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Dépenses du jour
    row += 1
    ws[f'A{row}'] = "Dépenses"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="DC3545")
    row += 1
    
    headers_depenses = ["ID", "Description", "Montant", "Date", "Utilisateur"]
    for col, header in enumerate(headers_depenses, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        cell.border = border
        cell.alignment = center_alignment
    
    row += 1
    depenses_du_jour = Depense.objects.filter(date_depense=today).order_by('-date_depense')
    for depense in depenses_du_jour:
        ws.cell(row=row, column=1, value=depense.id)
        ws.cell(row=row, column=2, value=depense.description)
        ws.cell(row=row, column=3, value=f"{depense.montant:,.0f}")
        ws.cell(row=row, column=4, value=depense.date_depense.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=5, value=depense.utilisateur.login if depense.utilisateur else "N/A")
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Ajuster la largeur des colonnes
    column_widths = [20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=dashboard_restaurant_{today.strftime("%Y%m%d")}.xlsx'
    
    # Sauvegarder le workbook dans la réponse
    wb.save(response)
    
    return response
